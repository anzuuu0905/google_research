from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
import time
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime
import os
from PIL import Image
import io

def optimize_screenshot(png_data, max_width=1024):
    image = Image.open(io.BytesIO(png_data))
    width_percent = max_width / float(image.size[0])
    new_height = int(float(image.size[1]) * width_percent)
    image = image.resize((max_width, new_height), Image.Resampling.LANCZOS)
    output = io.BytesIO()
    image.save(output, format='JPEG', quality=60, optimize=True)
    return output.getvalue()

def analyze_multiple_sites(urls):
    chrome_options = Options()
    chrome_options.add_argument('--headless')
    chrome_options.add_argument('--window-size=1024,768')

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=chrome_options)

    wb = Workbook()
    ws = wb.active
    ws.title = "サイト解析結果"

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    screenshot_dir = f'screenshots_{timestamp}'
    os.makedirs(screenshot_dir, exist_ok=True)

    try:
        for site_index, url in enumerate(urls):
            print(f"\nサイト {site_index + 1} の解析中: {url}")

            col_offset = site_index * 4
            headers = [f'タグ_{site_index+1}', f'レベル_{site_index+1}',
                      f'テキスト_{site_index+1}', f'URL_{site_index+1}']

            for i, header in enumerate(headers):
                cell = ws.cell(row=1, column=col_offset + i + 1)
                cell.value = header
                cell.font = Font(bold=True)

            try:
                driver.get(url)
                time.sleep(3)

                # スクリーンショット取得と最適化
                total_height = driver.execute_script("return document.body.scrollHeight")
                driver.set_window_size(1024, total_height)
                screenshot = driver.get_screenshot_as_png()
                optimized_screenshot = optimize_screenshot(screenshot)

                screenshot_path = os.path.join(screenshot_dir, f'site_{site_index+1}.jpg')
                with open(screenshot_path, 'wb') as f:
                    f.write(optimized_screenshot)
                print(f"スクリーンショット保存: {screenshot_path}")

                elements = driver.find_elements(By.XPATH, "//*[self::h1 or self::h2 or self::h3 or self::a]")
                current_level = 0
                seen_urls = set()
                row = 2

                for element in elements:
                    try:
                        tag_name = element.tag_name
                        text = ' '.join(element.text.strip().split())

                        if tag_name.startswith('h'):
                            level = int(tag_name[1])
                            current_level = level

                            if not text:
                                img_elements = element.find_elements(By.TAG_NAME, "img")
                                for img in img_elements:
                                    alt_text = img.get_attribute("alt")
                                    if alt_text:
                                        text = alt_text
                                        break

                            if text:
                                ws.cell(row=row, column=col_offset + 1, value=f'H{level}')
                                ws.cell(row=row, column=col_offset + 2, value=level)
                                ws.cell(row=row, column=col_offset + 3, value=text)
                                ws.cell(row=row, column=col_offset + 4, value='')
                                row += 1

                        elif tag_name == 'a':
                            href = element.get_attribute("href")
                            if not text:
                                img_elements = element.find_elements(By.TAG_NAME, "img")
                                for img in img_elements:
                                    alt_text = img.get_attribute("alt")
                                    if alt_text:
                                        text = alt_text
                                        break

                            if href and text and href not in seen_urls:
                                seen_urls.add(href)
                                ws.cell(row=row, column=col_offset + 1, value='a')
                                ws.cell(row=row, column=col_offset + 2, value=current_level)
                                ws.cell(row=row, column=col_offset + 3, value=text)
                                ws.cell(row=row, column=col_offset + 4, value=href)
                                row += 1

                    except Exception as e:
                        continue

            except Exception as e:
                print(f"Error processing URL {url}: {str(e)}")
                continue

        # 列幅の自動調整
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        excel_filename = f'multiple_sites_analysis_{timestamp}.xlsx'
        wb.save(excel_filename)
        print(f"\nExcelファイル '{excel_filename}' を保存しました。")

    finally:
        driver.quit()

def google_search_and_analyze(keyword):
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service)

    try:
        print(f"\nGoogle検索を実行中: {keyword}")
        driver.get("https://www.google.com")
        search_box = driver.find_element(By.NAME, "q")
        search_box.send_keys(keyword)
        search_box.send_keys(Keys.RETURN)
        time.sleep(2)

        search_results = driver.find_elements(By.CSS_SELECTOR, ".tF2Cxc")
        urls = []

        print("\n検索結果からURLを取得中...")
        for i, result in enumerate(search_results[:10], 1):
            try:
                url = result.find_element(By.CSS_SELECTOR, "a").get_attribute("href")
                title = result.find_element(By.CSS_SELECTOR, "h3").text
                if url and not url.startswith("https://www.google.com"):
                    urls.append(url)
                    print(f"{i}. タイトル: {title}")
                    print(f"   URL: {url}\n")
            except:
                continue

        driver.quit()
        print(f"\n取得した総検索結果: {len(urls)}件")

        if urls:
            print("\nサイトの解析を開始します...")
            analyze_multiple_sites(urls)
        else:
            print("解析可能なURLが見つかりませんでした。")

    except Exception as e:
        print(f"エラーが発生しました: {str(e)}")
        driver.quit()

if __name__ == "__main__":
    keyword = input("検索キーワードを入力してください: ")
    google_search_and_analyze(keyword)
